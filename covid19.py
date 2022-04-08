import requests
import json
import datetime
from openpyxl import load_workbook
import copy
from SPARQLWrapper import SPARQLWrapper, JSON
import pandas as pd
from iso3166 import countries
from sqlalchemy import create_engine
from bs4 import BeautifulSoup
from urllib.request import urlopen
import time
import sys
import os
from dotenv import load_dotenv
import ast



# ----------------------------------------------------------------------------------
# Functions
#

def get_newest_mortality_link():
    url = 'https://www.bfs.admin.ch/bfs/de/home/statistiken/gesundheit/gesundheitszustand/sterblichkeit-todesursachen' \
          '.assetdetail.16006453.html '
    page = requests.get(url, timeout=20)
    soup = BeautifulSoup(page.content, "html.parser")  # converts the page content into a beautifulsoup object
    new_url = f"https://www.bfs.admin.ch{soup.find('div', {'class': 'alert bg-success glyphicon-refresh text-success'}).find('a', href=True)['href']} "
    page = requests.get(new_url, timeout=20)
    soup = BeautifulSoup(page.content, "html.parser")  # converts the page content into a beautifulsoup object
    # new_link = f"https://www.bfs.admin.ch{soup.findAll('a', {'class': 'icon icon--before icon--doc js-ga-bfs-download-event'})[0]['href']}"
    new_link = f"{soup.findAll('a', {'class': 'icon icon--before icon--doc js-ga-bfs-download-event'})[0]['href']}"
    return new_link


def wikidata_get_population_all_countries():
    sparql = SPARQLWrapper("https://query.wikidata.org/sparql",
                           agent='Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_5) AppleWebKit/537.36 (KHTML, '
                                 'like Gecko) Chrome/50.0.2661.102 Safari/537.36')
    sparql.setQuery("""
    SELECT ?country ?countryLabel ?population ?countrycode{
    ?country wdt:P1082 ?population .
    ?country wdt:P298 ?countrycode.
    SERVICE wikibase:label { bd:serviceParam wikibase:language "en" }
    {
        SELECT ?country  WHERE {
        ?city wdt:P31 wd:Q515 .
        ?city wdt:P17 ?country .
        } GROUP BY ?country 
    }
    } """)
    sparql.setReturnFormat(JSON)
    results = sparql.query().convert()
    wikidata_pop = results["results"]["bindings"]

    _country_code = []
    _country_label = []
    _country_population = []
    for item in wikidata_pop:
        _country_code.append(item['countrycode']['value'])
        _country_label.append(item['countryLabel']['value'])
        _country_population.append(item['population']['value'])
    df_pop = pd.DataFrame(zip(_country_code,
                              _country_label,
                              _country_population), columns=['countrycode', 'countrylabel', 'population'])
    df_pop = df_pop.drop_duplicates(subset=['countrycode'])
    df_pop['population'] = pd.to_numeric(df_pop['population'])
    return df_pop


def is_internet_on():
    try:
        urlopen('https://www.google.com/', timeout=20)
        return True
    except:
        return False


def add_bfs_mortality(engine, table):
    ##########################################
    # add mortality 2010 ... 2019
    #
    #
    dateparse = lambda x: datetime.datetime.strptime(x, '%d.%m.%Y')
    url = 'https://www.bfs.admin.ch/bfsstatic/dam/assets/12607336/master'
    myfile = requests.get(url)
    open('ts-d-14.03.04.03-wr_ZR.csv', 'wb').write(myfile.content)
    df1 = pd.read_csv('ts-d-14.03.04.03-wr_ZR.csv', sep=';', nrows=1042, parse_dates=['Endend'], date_parser=dateparse,
                      na_values='           .')
    df1 = df1.rename(
        columns={'Alter': 'Age', 'Endend': 'Date', 'Anzahl_Todesfalle': 'AnzTF_HR', 'Kalenderwoche': 'Woche'})
    # store in sql later

    ##########################################
    # add mortality 2020 + 2021
    #
    #
    dateparse = lambda x: datetime.datetime.strptime(x, '%d.%m.%Y')
    url = get_newest_mortality_link()
    myfile = requests.get(url)
    open('ts-d-14.03.04.03-wr.csv', 'wb').write(myfile.content)
    df2 = pd.read_csv('ts-d-14.03.04.03-wr.csv', sep=';',nrows=314, parse_dates=['endend'], date_parser=dateparse,
                      na_values='           .')
    df2 = df2.rename(columns={'Alter': 'Age', 'endend': 'Date', 'Jahr': 'KJ'})

    # concatenate 2010 ... 2019 and 2020/2021
    df = pd.concat([df1, df2], ignore_index=True)
    df['Age'] = df['Age'].str.strip()

    df['AnzTF_HRfilt'] = df.sort_values('Date').groupby('Age')['AnzTF_HR'].rolling(12).mean().reset_index(0, drop=True)
    with engine.connect() as con:
        df.to_sql(name=table, con=con, if_exists='replace', index=True)
    engine.dispose()
    print(f"insert done into: {table}")
    return df


def add_bfs_mortalityrate(engine, table):
    ##########################################
    # add Todesfälle pro Jahr und Sterblichkeit
    #
    #
    dateparse = lambda x: datetime.datetime.strptime(x, '%Y')
    url = 'https://www.pxweb.bfs.admin.ch/sq/c9604a82-ad70-4a42-8409-10cb0a9e9e4c'
    myfile = requests.get(url)
    open('px-x-0102020206_111.csv', 'wb').write(myfile.content)
    df = pd.read_csv('px-x-0102020206_111.csv', parse_dates=['Jahr'], date_parser=dateparse, sep=';', encoding="cp1252",
                     na_values='...')
    df = df.rename(
        columns={'Todesfälle - Total': 'Deaths', 'Todesfälle je 1000 Einwohner': 'DeathsPer1k', 'Jahr': 'Date'})
    with engine.connect() as con:
        df.to_sql(name=table, con=con, if_exists='replace', index=True)
    engine.dispose()
    print(f"insert done into: {table}")
    return df


def add_bfs_population(engine, table):
    ##########################################
    # add population CH
    #
    #
    dateparse = lambda x: datetime.datetime.strptime(x, '%Y')
    url = 'https://www.pxweb.bfs.admin.ch/sq/1fbb332f-952d-41ad-9b9a-0abd8d98183d'
    myfile = requests.get(url)
    open('px-x-0102030000_101.csv', 'wb').write(myfile.content)
    df = pd.read_csv('px-x-0102030000_101.csv', parse_dates=['Jahr'], date_parser=dateparse, sep=';')
    df = df.rename(columns={'Geschlecht - Total': 'Population', 'Jahr': 'Date'})
    with engine.connect() as con:
        df.to_sql(name=table, con=con, if_exists='replace', index=True)
    engine.dispose()
    print(f"insert done into: {table}")
    return df


def add_openzh_covid19_data(engine, table):
    ##########################################
    # add data from openZH csv
    # used for icu, vent, hosp
    #
    url = 'https://raw.githubusercontent.com/openZH/covid_19/master/COVID19_Fallzahlen_CH_total_v2.csv'
    myfile = requests.get(url)
    open('COVID19_Fallzahlen_CH_total_v2.csv', 'wb').write(myfile.content)
    df = pd.read_csv('COVID19_Fallzahlen_CH_total_v2.csv')
    df['date'] = pd.to_datetime(df['date'])
    df['new_conf'] = df.sort_values('date').groupby('abbreviation_canton_and_fl')['ncumul_conf'].diff().fillna(0)

    data = {
        'abbreviation_canton_and_fl': ['ZH', 'BE', 'LU', 'UR', 'SZ', 'OW', 'NW', 'GL', 'ZG', 'FR', 'SO', 'BS', 'BL',
                                       'SH',
                                       'AR', 'AI', 'SG', 'GR', 'AG', 'TG', 'TI', 'VD', 'VS', 'NE', 'GE', 'JU', 'FL'],
        'canton_population': [1521.0, 1035.0, 409.6, 36.4, 159.2, 37.8, 43.2, 40.4, 126.8, 318.7, 273.2, 194.8, 288.1,
                              82.0,
                              55.2, 16.1, 507.7, 198.4, 678.2, 276.5, 353.3, 799.1, 344.0, 176.9, 499.5, 73.4, 38.7]
    }
    df2 = pd.DataFrame(data, columns=['abbreviation_canton_and_fl', 'canton_population'])
    df2.canton_population = df2.canton_population * 1000

    df['new_confper100k'] = 100e3 / df.abbreviation_canton_and_fl.map(
        df2.set_index('abbreviation_canton_and_fl').canton_population) * df.new_conf
    df['current_hospper100k'] = 100e3 / df.abbreviation_canton_and_fl.map(
        df2.set_index('abbreviation_canton_and_fl').canton_population) * df.current_hosp
    df['current_icuper100k'] = 100e3 / df.abbreviation_canton_and_fl.map(
        df2.set_index('abbreviation_canton_and_fl').canton_population) * df.current_icu
    df['current_ventper100k'] = 100e3 / df.abbreviation_canton_and_fl.map(
        df2.set_index('abbreviation_canton_and_fl').canton_population) * df.current_vent
    df['new_confper100kfilt7d'] = df.sort_values('date').groupby('abbreviation_canton_and_fl')[
        'new_confper100k'].rolling(
        7).mean().reset_index(0, drop=True)

    df['new_hospfilt7d'] = df.sort_values('date').groupby('abbreviation_canton_and_fl')['new_hosp'].rolling(
        7).mean().reset_index(0, drop=True)

    df['current_hospfilt7d'] = df.sort_values('date').groupby('abbreviation_canton_and_fl')['current_hosp'].rolling(
        7).mean().reset_index(0, drop=True)
    df['current_hospInterp'] = df.sort_values('date').groupby('abbreviation_canton_and_fl')['current_hosp'].apply(
        lambda group: group.interpolate(method='linear', limit_direction='forward'))
    df['current_hospInterpfilt7d'] = df.sort_values('date').groupby('abbreviation_canton_and_fl')[
        'current_hospInterp'].rolling(7).mean().reset_index(0, drop=True)
    df['current_hospInterpper100k'] = 100e3 / df.abbreviation_canton_and_fl.map(
        df2.set_index('abbreviation_canton_and_fl').canton_population) * df.current_hospInterp

    df['new_hosp2'] = df.sort_values('date').groupby('abbreviation_canton_and_fl')['current_hospInterp'].diff().fillna(
        0)
    df['new_hosp2filt7d'] = df.sort_values('date').groupby('abbreviation_canton_and_fl')['new_hosp2'].rolling(
        7).mean().reset_index(0, drop=True)

    df['current_icufilt7d'] = df.sort_values('date').groupby('abbreviation_canton_and_fl')['current_icu'].rolling(
        7).mean().reset_index(0, drop=True)
    df['current_icuInterp'] = df.sort_values('date').groupby('abbreviation_canton_and_fl')['current_icu'].apply(
        lambda group: group.interpolate(method='linear', limit_direction='forward'))
    df['new_icu2'] = df.sort_values('date').groupby('abbreviation_canton_and_fl')['current_icuInterp'].diff().fillna(0)
    df['new_icu2filt7d'] = df.sort_values('date').groupby('abbreviation_canton_and_fl')['new_icu2'].rolling(
        7).mean().reset_index(0, drop=True)

    df['current_ventfilt7d'] = df.sort_values('date').groupby('abbreviation_canton_and_fl')['current_vent'].rolling(
        7).mean().reset_index(0, drop=True)
    df['current_ventInterp'] = df.sort_values('date').groupby('abbreviation_canton_and_fl')['current_vent'].apply(
        lambda group: group.interpolate(method='linear', limit_direction='forward'))
    df['new_vent2'] = df.sort_values('date').groupby('abbreviation_canton_and_fl')['current_ventInterp'].diff().fillna(
        0)
    df['new_vent2filt7d'] = df.sort_values('date').groupby('abbreviation_canton_and_fl')['new_vent2'].rolling(
        7).mean().reset_index(0, drop=True)

    df['new_conffilt7d'] = df.sort_values('date').groupby('abbreviation_canton_and_fl')['new_conf'].rolling(
        7).mean().reset_index(0, drop=True)
    with engine.connect() as con:
        df.to_sql(name=table, con=con, if_exists='replace', index=True)
    engine.dispose()
    print(f"insert done into: {table}")
    return df


def add_owid_covid19_data(engine, table, config_data):
    ##########################################
    # add data from our world in data excel
    # used for tested all countries  ~12MB
    #
    #url = 'https://covid.ourworldindata.org/data/owid-covid-data.xlsx'
    url = 'https://covid.ourworldindata.org/data/owid-covid-data.csv'
    myfile = requests.get(url)
    #open('owid-covid-data.xlsx', 'wb').write(myfile.content)
    open('owid-covid-data.csv', 'wb').write(myfile.content)
    df = pd.read_csv(r'owid-covid-data.csv', sep=',', low_memory=False)

    country_code_three_digits = []
    for country_two_digit in config_data['countries_of_interest']:
        country_code_three_digits.append(countries.get(country_two_digit)[2])

    df = df.loc[df['iso_code'].isin(country_code_three_digits)]
    with engine.connect() as con:
        df.to_sql(name=table, con=con, if_exists='replace', index=True)
    engine.dispose()
    print(f"insert done into: {table}")


def add_covid19_tracker_data(engine, table, config_data, df_pop, add_SE2=True):
    ##########################################
    # get data from coronavirus-tracker API
    #
    url3 = "https://cvtapi.nl/all"
    url2 = "https://coronavirus-tracker-api.herokuapp.com/all"
    url1 = "https://covid-tracker-us.herokuapp.com/all"
    urls = (url1, url2, url3)

    valid_response = False
    response = []
    for url in urls:
        response = requests.get(url)
        if response.status_code == 200:
            valid_response = True
            break
        else:
            continue

    if not valid_response:
        print('Could not get data from the coronavirus-tracker API')
        return -1

    data = response.text
    parsed = json.loads(data)
    nr_of_countries = len(parsed["deaths"]["locations"])  # + 1  # x Countries +1 for sweden

    # To read out the array index
    country_code_two_digits = {}
    country_code_three_digits = {}
    for country in config_data["countries_of_interest"]:
        country_code_two_digits[country] = 0
        country_code_three_digits[countries.get(country)[2]] = 0

    # -------------------------------------------------------------------
    # Get the magic country code number out of the parsed json object.
    #
    for i in range(0, nr_of_countries - 2):
        if parsed["deaths"]["locations"][i]["country_code"] in country_code_two_digits and \
                parsed["deaths"]["locations"][i][
                    "province"] == "":
            country_code_two_digits[parsed["deaths"]["locations"][i]["country_code"]] = i
            country_code_three_digits[countries.get(parsed["deaths"]["locations"][i]["country_code"])[
                2]] = i  # change from 2digit to three digits.

    if add_SE2:
        ##########################################
        # add SE2 from Excel to JSON
        # used for more accurate deaths time of sweden
        #
        country_code_three_digits['SE2'] = len(parsed["deaths"]["locations"])
        SE2_confirmed_json = copy.deepcopy(parsed["confirmed"]["locations"][country_code_three_digits["SWE"]])
        SE2_deaths_json = copy.deepcopy(parsed["deaths"]["locations"][country_code_three_digits["SWE"]])
        parsed["deaths"]["locations"].append(SE2_deaths_json)
        parsed["confirmed"]["locations"].append(SE2_confirmed_json)

        url = 'https://www.arcgis.com/sharing/rest/content/items/b5e7488e117749c19881cce45db13f7e/data'
        myfile = requests.get(url, allow_redirects=True)
        open('Folkhalsomyndigheten_Covid19.xlsx', 'wb').write(myfile.content)
        wb = load_workbook(filename='Folkhalsomyndigheten_Covid19.xlsx')
        ws = wb['Antal avlidna per dag']

        SE2_DateDeaths = []
        for i in range(2, ws.max_row):
            if isinstance(ws.cell(row=i, column=1).value, datetime.date):
                SE2_DateDeaths.append(ws.cell(row=i, column=1).value)
            else:
                break

        SE2_Deaths = []
        for i in range(2, ws.max_row):
            if isinstance(ws.cell(row=i, column=1).value, datetime.date):
                if i >= 3:
                    SE2_Deaths.append(ws.cell(row=i, column=2).value + SE2_Deaths[i - 3])
                else:
                    SE2_Deaths.append(ws.cell(row=i, column=2).value)
            else:
                break

        k = 0
        for key, value in parsed["deaths"]["locations"][country_code_three_digits["SE2"]]["history"].items():
            if k >= len(SE2_DateDeaths):
                parsed["deaths"]["locations"][country_code_three_digits["SE2"]]["history"][key] = SE2_Deaths[k - 1]
            else:
                if datetime.datetime.strptime(key, "%m/%d/%y").date() == SE2_DateDeaths[k].date():
                    parsed["deaths"]["locations"][country_code_three_digits["SE2"]]["history"][key] = SE2_Deaths[k]
                    k += 1
                else:
                    parsed["deaths"]["locations"][country_code_three_digits["SE2"]]["history"][key] = 0

    df = pd.DataFrame()
    for three_digit, value_three_digit in country_code_three_digits.items():
        df = df.append(pd.DataFrame(
            zip(list(parsed["deaths"]["locations"][value_three_digit]["history"].keys()),
                list([three_digit for _x in range(len(
                    parsed["deaths"]["locations"][value_three_digit]["history"].items()))]),
                list([countries.get(parsed["confirmed"]["locations"][value_three_digit]["country_code"])[2] for _x in
                      range(len(
                          parsed["deaths"]["locations"][value_three_digit]["history"].items()))]),
                list(parsed["confirmed"]["locations"][value_three_digit]["history"].values()),
                list(parsed["deaths"]["locations"][value_three_digit]["history"].values())),
            columns=['date', 'countrycode', 'country_code_for_pop', 'confirmed', 'deaths']),
            ignore_index=True)
    df['date'] = pd.to_datetime(df['date'])

    df['confirmedperpop'] = 1e6 / df['country_code_for_pop'].map(df_pop.set_index('countrycode')['population']) * df[
        'confirmed']
    df['newconfirmed'] = df.sort_values('date').groupby('countrycode')['confirmed'].diff().fillna(0)
    df['newconfirmedperpop'] = 1e6 / df['country_code_for_pop'].map(df_pop.set_index('countrycode')['population']) * df[
        'newconfirmed']
    df['newconfirmedfilt3d'] = df.sort_values('date').groupby('countrycode')['newconfirmed'].rolling(
        3).mean().reset_index(0, drop=True)
    df['newconfirmedfilt7d'] = df.sort_values('date').groupby('countrycode')['newconfirmed'].rolling(
        7).mean().reset_index(0, drop=True)
    df['newconfirmedfilt14d'] = df.sort_values('date').groupby('countrycode')['newconfirmed'].rolling(
        14).mean().reset_index(0, drop=True)
    df['newconfirmedperpopfilt3d'] = df.sort_values('date').groupby('countrycode')['newconfirmedperpop'].rolling(
        3).mean().reset_index(0, drop=True)
    df['newconfirmedperpopfilt7d'] = df.sort_values('date').groupby('countrycode')['newconfirmedperpop'].rolling(
        7).mean().reset_index(0, drop=True)
    df['newconfirmedperpopfilt14d'] = df.sort_values('date').groupby('countrycode')['newconfirmedperpop'].rolling(
        14).mean().reset_index(0, drop=True)

    df['deathsperpop'] = 1e6 / df['country_code_for_pop'].map(df_pop.set_index('countrycode')['population']) * df[
        'deaths']
    df['newdeaths'] = df.sort_values('date').groupby('countrycode')['deaths'].diff().fillna(0)
    df['newdeathsperpop'] = 1e6 / df['country_code_for_pop'].map(df_pop.set_index('countrycode')['population']) * df[
        'newdeaths']
    df['newdeathsfilt3d'] = df.sort_values('date').groupby('countrycode')['newdeaths'].rolling(3).mean().reset_index(0,
                                                                                                                     drop=True)
    df['newdeathsfilt7d'] = df.sort_values('date').groupby('countrycode')['newdeaths'].rolling(7).mean().reset_index(0,
                                                                                                                     drop=True)
    df['newdeathsfilt14d'] = df.sort_values('date').groupby('countrycode')['newdeaths'].rolling(14).mean().reset_index(
        0,
        drop=True)
    df['newdeathsperpopfilt3d'] = df.sort_values('date').groupby('countrycode')['newdeathsperpop'].rolling(
        3).mean().reset_index(0, drop=True)
    df['newdeathsperpopfilt7d'] = df.sort_values('date').groupby('countrycode')['newdeathsperpop'].rolling(
        7).mean().reset_index(0, drop=True)
    df['newdeathsperpopfilt14d'] = df.sort_values('date').groupby('countrycode')['newdeathsperpop'].rolling(
        14).mean().reset_index(0, drop=True)
    del df['country_code_for_pop']  # was only used for calculation
    with engine.connect() as con:
        df.to_sql(name=table, con=con, if_exists='replace', index=True)
    engine.dispose()
    print(f"insert done into: {table}")
    return df


def add_bag_tested(engine, table):
    ##########################################
    # add tested CH from Excel
    # not really needed anymore (owid)
    #
    url = 'https://www.bag.admin.ch/dam/bag/de/dokumente/mt/k-und-i/aktuelle-ausbrueche-pandemien/2019-nCoV/covid-19-basisdaten-labortests.xlsx.download.xlsx/Dashboard_3_COVID19_labtests_positivity.xlsx'
    myfile = requests.get(url)
    open('Dashboard_3_COVID19_labtests_positivity.xlsx', 'wb').write(myfile.content)
    df = pd.read_excel(r'Dashboard_3_COVID19_labtests_positivity.xlsx')
    del df['Replikation_dt']
    df_new = df.pivot_table('Number_of_tests', ['Datum'], 'Outcome_tests')
    df_new.reset_index(drop=False, inplace=True)
    df_new.reindex(['date', 'Negative', 'Positive'], axis=1)
    df_new = df_new.rename(columns={'Outcome_tests': 'Index', 'Datum': 'date', 'Negative': 'newtestednegative',
                                    'Positive': 'newtestedpositive'})
    df_new['newtested'] = df_new['newtestednegative'] + df_new['newtestedpositive']
    df_new['newtestedfilt7d'] = df_new.sort_values('date')['newtested'].rolling(7).mean().reset_index(0, drop=True)
    df_new['newtestednegativefilt7d'] = df_new.sort_values('date')['newtestednegative'].rolling(7).mean().reset_index(0,
                                                                                                                      drop=True)
    df_new['newtestedpositivefilt7d'] = df_new.sort_values('date')['newtestedpositive'].rolling(7).mean().reset_index(0,
                                                                                                                      drop=True)
    df.fillna(0)
    with engine.connect() as con:
        df_new.to_sql(name=table, con=con, if_exists='replace', index=True)
    engine.dispose()
    print(f"insert done into: {table}")


def add_country_info(engine, table, df_pop, config_data):
    ##########################################
    # Additional Information
    #
    country_code_three_digits = []
    for country_two_digit in config_data['countries_of_interest']:
        country_code_three_digits.append(countries.get(country_two_digit)[2])
    df_country = pd.DataFrame(country_code_three_digits,
                              columns=['countrycode'])
    df_new = df_pop[df_pop['countrycode'].isin(df_country['countrycode'])]
    df_new.reset_index(drop=True, inplace=True)
    with engine.connect() as con:
        df_new.to_sql(name=table, con=con, if_exists='replace', index=True)
    engine.dispose()
    print(f"insert done into: {table}")


def read_json_cfg_file(file_path):
    with open(file_path, "r") as f:
        return json.load(f)


def log(engine, status):
    ##########################################
    # add entry in excel file
    #

    # this approach does not work for heroku, because the files will be not persistent
    #df_lastupdated = pd.read_csv(r'Last_Updated.csv', sep=';')

    # access DB to get last inserts. Other alternative would be to just add
    with engine.connect() as con:
        df_lastupdated = pd.read_sql('general_info', con=con, index_col=['index'])
    engine.dispose()
    df_lastupdated['last_update_time'] = pd.to_datetime(df_lastupdated['last_update_time'])
    d = {'last_update_time': [datetime.datetime.now().replace(microsecond=0)], 'status': status}
    df = pd.DataFrame(data=d)
    df_lastupdated = df_lastupdated.append(df, ignore_index=True)
    df_lastupdated.to_csv('Last_Updated.csv', index=False, sep=';')
    with engine.connect() as con:
        df_lastupdated.to_sql(name='general_info', con=con, if_exists='replace', index=True)
    engine.dispose()
    return df_lastupdated


def retry(fun, max_tries=10):
    for i in range(max_tries):
        try:
            time.sleep(0.3)
            fun()
            break
        except Exception:
            continue


# end of functins ------------------------------------------------------------------


def main():
    ##########################################
    # check internet connection
    #
    print("started")
    if not is_internet_on():
        print('The Internet connectino is gone, please reconnect and execute again')
        return -1

    load_dotenv()
    config_data = {}
    config_data["mysql-engine"] = os.getenv('mysql-engine')
    config_data["countries_of_interest"] = ast.literal_eval(os.getenv('countries_of_interest'))
    config_data["datasources"] = dict(ast.literal_eval(os.getenv('datasources')))

    engine = create_engine(config_data["mysql-engine"])
    sys.stdout.flush()

    log(engine, "started")

    ##########################################
    # get population of All countries
    #
    df_pop = wikidata_get_population_all_countries()

    ##########################################
    # add mortality from 2010 ... 2021
    #
    if config_data["datasources"]["bfs_mortality"]:
        add_bfs_mortality(engine, "bfs_covid19")

    ##########################################
    # add population from 1860 ... 2019
    #
    if config_data["datasources"]["bfs_population"]:
        add_bfs_population(engine, "bfs2_covid19")

    ##########################################
    # add mortalityrate from 1803 ... 2019
    #
    if config_data["datasources"]["bfs_mortalityrate"]:
        add_bfs_mortalityrate(engine, "bfs3_covid19")

    ##########################################
    # add tested from bag (also in owid data)
    #
    if config_data["datasources"]["bag_tested"]:
        add_bag_tested(engine, "bag_tested")

    ##########################################
    # add openzh data (cantons)
    #
    if config_data["datasources"]["openzh_covid19"]:
        retry(add_openzh_covid19_data(engine, "openzh_covid19"), 5)

    ##########################################
    # get data from coronavirus-tracker API
    #
    if config_data["datasources"]["covid_tracker"]:
        retry(add_covid19_tracker_data(engine, "covid19", config_data, df_pop, add_SE2=True), 5)

    add_country_info(engine, "country_info", df_pop, config_data)

    ##########################################
    # add owid covid19 dataset
    #
    if config_data["datasources"]["owid_covid19"]:
        retry(add_owid_covid19_data(engine, "owid_covid19", config_data), 5)

    log(engine, "finished")

    print("finished")


if __name__ == "__main__":
    main()
